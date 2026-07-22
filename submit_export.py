"""パズル型設計の内容を、代表の提出用『ガチャ設計シート』2種に流し込んでxlsxを生成。

- テンプレ①（マスター版・簡易）= templates/submit_master.xlsx（設計テンプレートシート）
- テンプレ②（v3.0・自動判定付き）= templates/submit_v3.xlsx（設計入力シート）

いずれも「入力セルに値を書く→数式(自動判定/公開前チェック/出現率等)はExcel/Sheetsで
自動再計算」方式。openpyxlは再計算しないので、開いた時に確定値が出る前提。
"""
from __future__ import annotations

import io
import os
from typing import List

import openpyxl

_TPL_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "templates")


def _active(rows: list) -> list:
    return [p for p in rows if (getattr(p, "name", "") or "").strip() or getattr(p, "count", 0)]


def _direct(p):
    d = getattr(p, "display_pt_direct", None)
    if d in (None, ""):
        return None
    try:
        return int(float(d))
    except (ValueError, TypeError):
        return None


def _markup(p):
    m = getattr(p, "markup", 0) or 0
    return m if m > 0 else None


def fill_v3(meta, rows) -> bytes:
    """v3.0（6シート・自動判定付き）の設計入力シートに流し込む。賞品は最大21行。"""
    wb = openpyxl.load_workbook(os.path.join(_TPL_DIR, "submit_v3.xlsx"))
    ws = wb["設計入力"]
    ws["D4"] = meta.title or "（無題）"
    ws["D5"] = meta.unit_price
    ws["D6"] = meta.total_tickets
    ws["D7"] = str(meta.limit_per_day).strip() or "なし"
    ws["D8"] = str(meta.limit_total).strip() or "なし"
    ws["D9"] = meta.cost_rate
    # 賞品テーブル 13〜33 の入力列(A〜I)をクリアしてから書く
    for r in range(13, 34):
        for col in "ABCDEFGHI":
            ws[f"{col}{r}"] = None
    for i, p in enumerate(_active(rows)[:21]):
        r = 13 + i
        ws[f"A{r}"] = p.rank
        ws[f"B{r}"] = p.name
        ws[f"C{r}"] = p.model_no or "-"
        ws[f"D{r}"] = p.count
        ws[f"E{r}"] = p.real_value
        ws[f"F{r}"] = p.shipping
        ws[f"G{r}"] = p.method
        ws[f"H{r}"] = _markup(p)
        ws[f"I{r}"] = _direct(p)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def fill_master(meta, rows) -> bytes:
    """マスター版（簡易パズルシート）の設計テンプレートシートに流し込む。賞品は最大371行。

    ※マスター版は受取方法列が無い簡易版。表示PT直接はR列、本数はL列(本数変更)。
    """
    wb = openpyxl.load_workbook(os.path.join(_TPL_DIR, "submit_master.xlsx"))
    ws = wb["設計テンプレート"]
    ws["B3"] = None                          # ガチャNo.は空（No連動でなく直接入力）
    ws["E3"] = meta.title or "（無題）"       # タイトル
    ws["C4"] = meta.unit_price                # 単価(上書き)
    ws["F4"] = meta.total_tickets             # 総口数(上書き)
    ws["B2"] = meta.charge_amount or 0        # 引く権利(課金額)
    # 賞品行 9〜379 の入力列をクリア（A賞ランク/B名/Cレア/H実価値/I倍率/K除外/L本数/R表示PT直接）
    for r in range(9, 380):
        for col in ("A", "B", "C", "H", "I", "K", "L", "R"):
            ws[f"{col}{r}"] = None
    for i, p in enumerate(_active(rows)[:371]):
        r = 9 + i
        ws[f"A{r}"] = p.rank
        ws[f"B{r}"] = p.name
        ws[f"C{r}"] = getattr(p, "model_no", "") or ""   # レア列は型番で代用（レア情報を持たないため）
        ws[f"H{r}"] = p.real_value
        ws[f"I{r}"] = _markup(p)
        ws[f"L{r}"] = p.count                              # 本数(変更)
        ws[f"R{r}"] = _direct(p)                           # 表示PT直接(任意)
        if getattr(p, "exclude", False):
            ws[f"K{r}"] = "除外"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
