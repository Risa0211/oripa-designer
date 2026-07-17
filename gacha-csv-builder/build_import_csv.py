#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
みんなのトレカ 管理画面 ガチャ用CSV 自動生成ツール
=====================================================

やること:
  ① マスターDB（カード原簿: 型番 → カード名/レアリティ/画像URL...）
  ② ガチャ設計シート（賞ごと: 型番/ランク/在庫/還元pt...）
  を「型番」で結合し、管理画面「商品→カード一覧→CSVのインポート」に
  そのまま流し込める A〜L 列のCSVを自動生成する。

  → これで「CSVに画像URLを手で照合して書く」作業がゼロになる。

前提（管理画面マニュアル P21-22 準拠）:
  出力CSVの列（英語ヘッダは変更禁止・並び順もこの通り）
    A URL                元サイト参照(非表示)
    B Title              カード名(ユーザー表示)
    C Description        社内メモ(非表示)
    D Price              参照価格(非表示)
    E Redemption Points  還元ポイント(ユーザー表示)
    F Image URL-src      画像URL(ユーザー表示) ← 保管庫のURL
    G Category           カードフォルダーの分類(事前作成必須)
    H Inventory          在庫数
    I Usage Limit        ラストワン賞の口数(通常空欄)
    J Video              当選演出動画名(動画DB登録名と完全一致)
    K Card Rank          1等/2等...
    L Badges             PSA10 等（複数は「,」区切り／原簿の想定に合わせ全角も可）

  ※ 実際の管理画面から落とせる「CSVサンプル」でヘッダ文字列を最終確定すること。
     ヘッダ文字列は columns.json で差し替え可能にしてある。

使い方:
  python3 build_import_csv.py \
      --master   master_db.csv \
      --design   gacha_design.csv \
      --out      import.csv

  # ヘッダ名を実サンプルに合わせたい場合:
  python3 build_import_csv.py ... --columns columns.json
"""

import argparse
import csv
import json
import re
import sys
import unicodedata
from pathlib import Path

import palette_lookup  # 演出カードの種別→パレットキー自動導出

# 設計テンプレ「設計入力」賞品テーブルの列ヘッダ → ビルダー内部キーの対応
DESIGN_XLSX_MAP = {
    "賞ランク": "ランク",
    "賞品名": "カード名",
    "型番": "型番",
    "口数": "在庫",
    "表示PT/枚\n(自動)": "還元ポイント",
    "表示PT/枚(自動)": "還元ポイント",
    "表示PT/枚": "還元ポイント",
    "実価値/枚\n(円)": "参照価格",   # D Price に使う（管理画面で必須の価格）
    "実価値/枚(円)": "参照価格",
    "実価値/枚": "参照価格",
    "実価値": "参照価格",
    "メモ": "社内メモ",
    # --- 演出カード（画像にpt/個数が焼き込まれた擬似カード）用 ---
    "種別": "種別",            # PSA10/パック/BOX/ポイント交換/なにかのカード/なにかAR・HR/福袋
    "演出種別": "種別",
    "表示pt": "表示pt",        # 画像に印字されているpt（PSA10/ポイント交換の自動特定に使う）
    "表示PT": "表示pt",
    "個数": "個数",            # パック×N の N
    "演出キー": "演出キー",      # パレットキーを直接指定（手動選択）
    "画像URL": "画像URL上書き",  # 画像URLを直接指定（最優先の手動選択）
    "画像URL上書き": "画像URL上書き",
}


def read_design_xlsx(path: str, sheet: str = "設計入力"):
    """自社のガチャ設計テンプレ(.xlsx)の『賞品テーブル』を直接読む。
    ヘッダ行(A列='賞ランク')を自動検出し、『合計』行 or 空行まで取得。"""
    import openpyxl  # 遅延import（CSV運用だけなら不要）
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames:
        sys.exit(f"[ERROR] シート'{sheet}'が無い。存在: {wb.sheetnames}")
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    # ヘッダ行検出
    hdr_i = None
    for i, r in enumerate(rows):
        if r and str(r[0]).strip() == "賞ランク":
            hdr_i = i
            break
    if hdr_i is None:
        sys.exit("[ERROR] 賞品テーブルのヘッダ(賞ランク)が見つからない")
    headers = [("" if c is None else str(c).strip()) for c in rows[hdr_i]]
    out = []
    for r in rows[hdr_i + 1:]:
        a = "" if r[0] is None else str(r[0]).strip()
        if a in ("", "合計"):
            break
        d = {}
        for c_i, h in enumerate(headers):
            key = DESIGN_XLSX_MAP.get(h)
            if key and c_i < len(r) and r[c_i] is not None:
                v = r[c_i]
                # 口数などの 1.0 → 1 整形
                if isinstance(v, float) and v.is_integer():
                    v = int(v)
                d[key] = str(v)
        if d:
            out.append(d)
    return out

# ---- 出力CSVのヘッダ（A〜L）。実サンプルで要確定。columns.json で上書き可。----
DEFAULT_HEADERS = [
    "URL",                 # A
    "Title",               # B
    "Description",         # C
    "Price",               # D
    "Redemption Points",   # E
    "Image URL-src",       # F
    "Category",            # G
    "Inventory",           # H
    "Usage Limit",         # I
    "Video",               # J
    "Card Rank",           # K
    "Badges",              # L
]


def norm_key(s: str) -> str:
    """型番の表記ゆれを吸収して照合キーにする。
    例: '201/165', '２０１/１６５', ' 201 / 165 ' → '201/165'
    ★DOPA原簿の型番は '{286/SM-P}' や '{015/034} [CP1]' のように波括弧やセットタグ付き
      → 角括弧の中身(セットタグ)を除去し、残った括弧文字も除去して核の型番で照合する。
    全角→半角、空白除去、大文字化。"""
    if s is None:
        return ""
    s = unicodedata.normalize("NFKC", str(s))          # 全角英数記号→半角
    s = re.sub(r"[\[【［].*?[\]】］]", "", s)             # [CP1] [その他] 等のセットタグを除去
    s = re.sub(r"[\{\}｛｝\[\]［］（）()【】]", "", s)     # 残った括弧文字を除去（{286/SM-P}→286/SM-P）
    s = s.replace(" ", "").replace("　", "").strip()
    return s.upper()


def read_csv_dict(path: str):
    p = Path(path)
    if not p.exists():
        sys.exit(f"[ERROR] ファイルが見つかりません: {path}")
    # utf-8-sig でBOM対応（ExcelやSheets書き出しのBOM込みを吸収）
    with p.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = [ {(k.strip() if k else k): (v if v is not None else "") for k, v in r.items()} for r in reader ]
    return rows


def get(row: dict, *names, default=""):
    """行から複数候補の列名でゆるく取得（日本語/英語ゆれ吸収）。"""
    for n in names:
        if n in row and str(row[n]).strip() != "":
            return str(row[n]).strip()
    return default


# 自社ドメイン（保管庫WP / 管理画面 / 業者倉庫S3）。これ以外(=競合dopa等)のURLはA列に出さない。
_OWN_DOMAINS = ("minnano-toreka.com", "minnano-toreca.com", "s3.minnano-toreca.com")


def _is_own_url(url: str) -> bool:
    u = (url or "").lower()
    if not u.startswith(("http://", "https://")):
        return True  # 相対/自由記述はそのまま許可（競合の絶対URLだけ弾く）
    return any(d in u for d in _OWN_DOMAINS)


def _name_key(s):
    """カード名の照合キー（全角半角/空白/括弧内注記のゆれを吸収）。"""
    s = unicodedata.normalize("NFKC", str(s or ""))
    s = s.replace(" ", "").replace("　", "")
    return s.strip().upper()


import re as _re
# 末尾のレア表記。裸(リーリエSR) / 括弧付き(リーリエ(SR)・リーリエ【SR】・リーリエ（SR）) 両対応。
_RAR_TOKEN = r"(SAR|SSR|CSR|CHR|UR|HR|SR|RRR|RR|ACE|AR|PROMO|プロモ)"
_RAR_SUFFIX = _re.compile(r"[（(\[【｛{]?" + _RAR_TOKEN + r"[）)\]】｝}]?$")


def _strip_rarity(name_key):
    """賞品名末尾のレア表記（例: アセロラのいたずらSR / コダック(AR)）を落として基底名にする。"""
    return _RAR_SUFFIX.sub("", name_key)


def _design_rarity(design_name):
    """賞品名末尾のレア表記を取り出す（無ければ ''）。照合の絞り込みに使う。"""
    m = _RAR_SUFFIX.search(_name_key(design_name))
    return m.group(1) if m else ""


def build(master_rows, design_rows, headers, generic_map=None, palette=None):
    generic_map = generic_map or {}
    # マスターを型番でインデックス化（型番は重複するのでリストで保持＝2段階照合用）
    type_index = {}
    name_index = {}
    for r in master_rows:
        k = norm_key(get(r, "型番", "kataban", "card_number", "number"))
        if k:
            type_index.setdefault(k, []).append(r)
        nk = _name_key(get(r, "カード名", "name", "title"))
        if nk:
            name_index.setdefault(nk, []).append(r)
    ndup = sum(1 for v in type_index.values() if len(v) > 1)
    if ndup:
        print(f"[INFO] マスターDBに型番重複 {ndup}種（→ 賞品名で自動特定）", file=sys.stderr)

    out_rows = []
    unmatched = []      # 保管庫に無い＝画像追加が必要
    ambiguous = []      # 名前で複数候補＝サムネ見て人が1枚選ぶ（picker行き）
    warnings = []

    def pal_row(key):
        return (palette or {}).get("by_key", {}).get(key)

    def cand_dict(c):
        return {"型番": get(c, "型番"), "レアリティ": get(c, "レアリティ", "rarity"),
                "カード名": get(c, "カード名"), "画像URL": get(c, "画像URL", "image_url", "Image URL-src")}

    for i, d in enumerate(design_rows, start=2):  # 2 = ヘッダ次の行番号(人が見やすいよう)
        raw_kata = get(d, "型番", "kataban", "card_number", "number")
        design_name = get(d, "カード名", "name", "title")
        key = norm_key(raw_kata)
        has_kata = bool(key) and raw_kata.strip() not in ("-", "ー", "―")

        url_override = get(d, "画像URL上書き", "画像URL", "image_url_override")
        enshutsu_key = get(d, "演出キー", "palette_key")
        shubetsu     = get(d, "種別", "演出種別")

        # --- 設計側（賞ごと）を先に取得（演出ptの導出に redeem を使う）---
        rank       = get(d, "ランク", "card_rank", "rank")          # 1等/2等...
        inventory  = get(d, "在庫", "口数", "inventory")
        redeem     = get(d, "還元ポイント", "redemption_points", "redeem")

        # 種別が未指定でも賞品名から演出種別を推定（テンプレに種別列が無い運用に対応）
        shu_from_name, kosu_from_name = palette_lookup.infer_shubetsu_from_name(design_name)
        eff_shubetsu = shubetsu or shu_from_name

        m = {}          # マスター行（実カードのとき埋まる）
        title = get(d, "タイトル上書き", "title_override") or design_name
        image_url = price = source_url = ""
        category = get(d, "カテゴリ", "category")

        # ===== 画像ソースの決定（優先順）=====
        # ① 画像URL直指定（最優先の手動選択）
        if url_override:
            image_url = url_override

        # ② 演出キー直指定（手動でパレットから選ぶ）
        elif enshutsu_key:
            p = pal_row(enshutsu_key)
            if p is None:
                unmatched.append({"row": i, "型番": raw_kata, "設計上の名前": design_name,
                                  "種別": f"演出キー'{enshutsu_key}'がパレットに無い"})
                continue
            image_url = p.get("画像URL", "")
            if not get(d, "タイトル上書き", "title_override"):
                title = design_name or (p.get("使用中の表記名", "").split(" / ")[0])

        # ③ 演出カード：種別（明示 or 名前から推定）＋表示pt/個数 でパレット導出
        elif eff_shubetsu:
            pt_hint = get(d, "表示pt", "表示PT") or redeem   # PSA10/ポイントは表示pt優先・無ければ還元pt
            kosu    = get(d, "個数") or (str(kosu_from_name) if kosu_from_name else "")
            pkey, why = palette_lookup.derive_key(palette, eff_shubetsu, pt_hint, kosu)
            if pkey is None:
                unmatched.append({"row": i, "型番": raw_kata, "設計上の名前": design_name,
                                  "種別": f"演出カード({eff_shubetsu})未解決: {why}"})
                continue
            p = pal_row(pkey)
            image_url = p.get("画像URL", "") if p else ""
            if not get(d, "タイトル上書き", "title_override"):
                title = design_name or ((p.get("使用中の表記名", "").split(" / ")[0]) if p else "")

        # ④ 実カード：★名前を主キーに照合。型番は「同名で絵柄が複数」時の絞り込みだけに使う。
        #   （型番が賞品名と食い違う場合は型番を信用せず、無関係な型番仲間は候補に出さない）
        else:
            nk = _name_key(design_name)
            # まず賞品名で照合（末尾レア表記を剥がした基底名も試す）
            cands = name_index.get(nk) or name_index.get(_strip_rarity(nk)) or []
            dr = _design_rarity(design_name)
            if dr and len(cands) > 1:
                rared = [c for c in cands if _name_key(get(c, "レアリティ", "rarity")) == dr]
                if rared:
                    cands = rared
            if cands:
                # 同名で絵柄が複数 → 型番があれば型番で1枚に絞る（型番＝絵柄の指定）
                if has_kata and len(cands) > 1:
                    both = [c for c in cands
                            if norm_key(get(c, "型番", "kataban", "card_number", "number")) == key]
                    if both:
                        cands = both
                    elif key in type_index:
                        # 型番は原簿にあるが別カード＝賞品名と食い違い
                        warnings.append(
                            f"設計 {i}行目「{design_name}」: 型番{raw_kata}は別カードの型番→無視し名前で照合")
                    else:
                        # ★再発検知: 型番が原簿に1件も無い＝表記ゆれ/未収録の可能性。手動送りの原因になる。
                        warnings.append(
                            f"設計 {i}行目「{design_name}」: 型番{raw_kata}が原簿に見つからず絞り込めず"
                            f"（表記ゆれ/未収録の可能性・要確認）")
            elif has_kata:
                # 賞品名がマスターに無い → 型番で引くが、候補名が賞品名と一致する時だけ採用。
                # 別カードの型番仲間（例: コダックに066/060=リーリエ等）は絶対に候補にしない。
                kcands = type_index.get(key, [])
                named = [c for c in kcands if _name_key(get(c, "カード名")) == nk]
                cands = named
                if not named and kcands:
                    warnings.append(
                        f"設計 {i}行目「{design_name}」: 型番{raw_kata}は別カード{len(kcands)}件に該当し賞品名と不一致→要追加扱い")

            if len(cands) == 1:
                m = cands[0]
                image_url  = get(m, "画像URL", "image_url", "image", "Image URL-src")
                price      = get(m, "参照価格", "price", "Price")
                source_url = get(m, "元URL", "source_url", "url", "URL")
                # A列(非表示の元サイト参照)に競合ドメインのURLを残さない（DOPA等の取込元を秘匿）
                if source_url and not _is_own_url(source_url):
                    source_url = ""
                category   = category or get(m, "カテゴリ", "category", "Category")
            elif len(cands) > 1:
                # 複数の絵柄が該当 → 勝手に決めず picker へ（型番を入れれば一意化できる旨を案内）
                ambiguous.append({"row": i, "ランク": rank, "設計上の名前": design_name,
                                  "還元pt": redeem, "候補": [cand_dict(c) for c in cands]})
                continue
            else:
                unmatched.append({"row": i, "型番": raw_kata, "設計上の名前": design_name,
                                  "種別": "実カード(保管庫に無し・要画像追加)"})
                continue
        usage_lim  = get(d, "ラストワン口数", "usage_limit")
        video      = get(d, "動画", "video") or get(m, "デフォルト動画", "default_video", "video")
        badges     = get(d, "バッジ", "badges") or get(m, "デフォルトバッジ", "default_badges", "badges")
        desc       = get(d, "社内メモ", "description")

        # --- 管理画面の必須項目を埋める（A URL / D Price / G Category は空だとインポート弾かれる）---
        # A URL: 元サイト参照は競合URLを出さない方針で空。必須なので自社の画像URLで代替（非表示・害なし）
        a_url = source_url or image_url
        # D Price: 原簿に参照価格が無い(DOPA)ので、設計の実価値/枚→無ければ還元ptで代替（数値必須）
        price = price or get(d, "参照価格", "実価値", "price", "Price") or redeem

        # --- 検証（ユーザー表示に関わる欠落は必ず知らせる）---
        if not image_url:
            warnings.append(f"設計 {i}行目 型番{key}: 画像URLが原簿に無い（保管庫へ画像投入＋原簿追記が必要）")
        if not rank:
            warnings.append(f"設計 {i}行目 型番{key}: ランク(1等/2等...)が未入力")
        if not redeem:
            warnings.append(f"設計 {i}行目 型番{key}: 還元ポイントが未入力")
        if not inventory:
            warnings.append(f"設計 {i}行目 型番{key}: 在庫数が未入力")
        if not category:
            warnings.append(f"設計 {i}行目「{design_name}」: カテゴリ(G)未設定→管理画面で弾かれます。①でカテゴリを指定してください")

        record = {
            "URL": a_url,
            "Title": title,
            "Description": desc,
            "Price": price,
            "Redemption Points": redeem,
            "Image URL-src": image_url,
            "Category": category,
            "Inventory": inventory,
            "Usage Limit": usage_lim,
            "Video": video,
            "Card Rank": rank,
            "Badges": badges,
        }
        # headers（実サンプルのヘッダ）順で並べる。role→実ヘッダ対応は headers の順序=A..L 固定。
        out_rows.append([record[role] for role in [
            "URL", "Title", "Description", "Price", "Redemption Points",
            "Image URL-src", "Category", "Inventory", "Usage Limit",
            "Video", "Card Rank", "Badges",
        ]])

    return out_rows, unmatched, warnings, ambiguous


def write_picker_html(path, ambiguous):
    """複数候補（同名で絵柄違い）を、サムネ付きで人が1枚選ぶための一覧HTML。"""
    esc = lambda s: (str(s or "")).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    parts = ["""<!doctype html><meta charset="utf-8"><title>候補から画像を選ぶ</title>
<style>body{font-family:system-ui,'Hiragino Sans',sans-serif;margin:24px;background:#faf9f7}
h1{font-size:18px}.hint{color:#555;font-size:13px;margin:6px 0 18px}
.row{background:#fff;border:1px solid #e5e2dd;border-radius:10px;padding:14px 16px;margin:14px 0}
.rhead{font-weight:700;margin-bottom:10px}.rhead .rk{color:#c0392b}
.cards{display:flex;flex-wrap:wrap;gap:14px}
.card{width:150px;border:1px solid #e0ddd7;border-radius:8px;padding:8px;text-align:center;background:#fff}
.card img{width:134px;height:auto;border-radius:4px;background:#f0efec}
.kata{font-weight:700;font-size:13px;margin-top:6px}.rar{color:#2c7;font-size:12px}
.copy{font-size:11px;color:#888;word-break:break-all;margin-top:4px}</style>
<h1>候補から画像を選ぶ（同名で絵柄が複数）</h1>
<div class="hint">各賞について正しい絵柄の【型番】を設計シートの型番列に入れて再実行すると、次回から自動で確定します。</div>"""]
    for a in ambiguous:
        parts.append(f'<div class="row"><div class="rhead"><span class="rk">{esc(a["ランク"])}</span> '
                     f'{esc(a["設計上の名前"])} <span style="color:#888;font-weight:400">'
                     f'（還元{esc(a["還元pt"])}pt・{len(a["候補"])}候補）</span></div><div class="cards">')
        for c in a["候補"]:
            parts.append(f'<div class="card"><img src="{esc(c["画像URL"])}" loading="lazy">'
                         f'<div class="kata">{esc(c["型番"])}</div>'
                         f'<div class="rar">{esc(c["レアリティ"])}</div>'
                         f'<div class="copy">{esc(c["カード名"])}</div></div>')
        parts.append("</div></div>")
    Path(path).write_text("\n".join(parts), encoding="utf-8")


def main():
    ap = argparse.ArgumentParser(description="みんなのトレカ ガチャCSV 生成")
    ap.add_argument("--master", required=True, help="マスターDB(カード原簿) CSV")
    ap.add_argument("--design", help="ガチャ設計シート CSV")
    ap.add_argument("--design-xlsx", help="自社ガチャ設計テンプレ .xlsx（賞品テーブルを直接読む）")
    ap.add_argument("--sheet", default="設計入力", help="--design-xlsx 内のシート名")
    ap.add_argument("--generic", help="型番=- の擬似カード用 汎用画像テーブルCSV(賞品名→画像URL)")
    ap.add_argument("--palette", nargs="*", default=["palette_pseudo.csv", "palette_extra.csv"],
                    help="演出カードのパレットCSV(複数可)。種別＋表示pt/個数から画像を自動導出")
    ap.add_argument("--out", default="import.csv", help="出力CSV(A〜L)")
    ap.add_argument("--columns", help="ヘッダ名上書きJSON(任意)")
    args = ap.parse_args()

    if not args.design and not args.design_xlsx:
        sys.exit("[ERROR] --design か --design-xlsx のどちらかが必要")

    headers = DEFAULT_HEADERS
    if args.columns:
        headers = json.loads(Path(args.columns).read_text(encoding="utf-8"))
        if len(headers) != 12:
            sys.exit("[ERROR] columns.json は12個のヘッダ配列にしてください(A〜L)")

    master_rows = read_csv_dict(args.master)
    design_rows = read_design_xlsx(args.design_xlsx, args.sheet) if args.design_xlsx else read_csv_dict(args.design)

    generic_map = {}
    if args.generic:
        for r in read_csv_dict(args.generic):
            name = get(r, "賞品名", "name")
            if name:
                generic_map[name] = r

    # 演出カード パレット（存在するファイルだけ読む）
    palette = palette_lookup.load_palette(*[p for p in (args.palette or []) if Path(p).exists()])

    out_rows, unmatched, warnings, ambiguous = build(
        master_rows, design_rows, headers, generic_map, palette)

    # 出力（BOM付きutf-8: 管理画面/Excel互換のため）
    with Path(args.out).open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerows(out_rows)

    # 未照合レポート（保管庫に無い＝画像追加が必要）
    if unmatched:
        rep = Path(args.out).with_name("unmatched.csv")
        with rep.open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=["row", "型番", "設計上の名前", "種別"], extrasaction="ignore")
            w.writeheader()
            w.writerows(unmatched)

    # 複数候補（picker）: サムネ付きHTML＋CSV
    picker_html = None
    if ambiguous:
        picker_html = Path(args.out).with_name("picker.html")
        write_picker_html(picker_html, ambiguous)
        pc = Path(args.out).with_name("ambiguous.csv")
        with pc.open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(["row", "ランク", "設計上の名前", "還元pt", "候補数", "候補(型番:レア …)"])
            for a in ambiguous:
                cs = " / ".join(f"{c['型番']}:{c['レアリティ']}" for c in a["候補"])
                w.writerow([a["row"], a["ランク"], a["設計上の名前"], a["還元pt"], len(a["候補"]), cs])

    # サマリ
    print("=" * 56)
    print(f"生成完了: {args.out}")
    print(f"  ① 確定してCSV出力       : {len(out_rows)}件")
    print(f"  ② 複数候補（要・画像選択）: {len(ambiguous)}件")
    print(f"  ③ 保管庫に無い（要・追加）: {len(unmatched)}件")
    if ambiguous:
        print(f"  → picker.html を開いて正しい絵柄の型番を設計シートに入れて再実行（次回から自動確定）")
        for a in ambiguous[:10]:
            print(f"     - {a['row']}行 {a['ランク']} {a['設計上の名前']}（{len(a['候補'])}候補）")
    if unmatched:
        print(f"  → unmatched.csv。保管庫に画像を追加（DOPA取込 or 手動アップ）してから再実行")
        for u in unmatched[:10]:
            print(f"     - {u['row']}行 {u['設計上の名前']}｜{u['種別']}")
    if warnings:
        print(f"  警告 {len(warnings)}件:")
        for wmsg in warnings[:12]:
            print(f"     ! {wmsg}")
    print("=" * 56)


if __name__ == "__main__":
    main()
